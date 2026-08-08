from __future__ import annotations

import json
import math
import os
import re
import tempfile
import unicodedata
from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook


MONTH_ORDER = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


@dataclass(frozen=True)
class SheetAudit:
    sheet: str
    rows_read: int
    rows_in_year: int
    max_week: int | None
    rows_latest_week: int
    duplicates: int = 0


@dataclass(frozen=True)
class BuildResult:
    year: int
    updated_to_week: int
    default_month: str
    directory_rows: int
    metric_rows: int
    ctc_dm_rows: int
    output: str
    audit_output: str


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _header_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _clean_text(value))
    return "".join(ch for ch in text.lower() if not unicodedata.combining(ch))


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = _clean_text(value).replace(",", "")
        try:
            number = float(text)
        except ValueError:
            return None
    return number if math.isfinite(number) else None


def parse_percent(value: Any) -> float | None:
    """Convierte '58.8 %' a 0.588 y conserva decimales ya normalizados."""
    if value is None or value == "":
        return None
    if isinstance(value, str):
        text = _clean_text(value)
        has_percent = "%" in text
        number = _number(text.replace("%", ""))
        if number is None:
            return None
        return number / 100 if has_percent else number
    return _number(value)


def normalize_ceco(value: Any, *, decimal_code: bool = False) -> str | None:
    number = _number(value)
    if number is None:
        text = _clean_text(value)
        digits = re.sub(r"\D", "", text)
        return digits or None
    if decimal_code and abs(number) < 1000:
        number *= 1000
    return str(int(round(number)))


def _rounded(value: Any, digits: int = 6) -> int | float | None:
    number = _number(value)
    if number is None:
        return None
    rounded = round(number, digits)
    return int(rounded) if rounded.is_integer() else rounded


def _required_sheet(workbook: Any, name: str) -> Any:
    if name not in workbook.sheetnames:
        raise ValueError(f"Falta la pestaña obligatoria: {name}")
    return workbook[name]


def _header_map(row: Iterable[Any]) -> dict[str, int]:
    return {_header_key(value): index for index, value in enumerate(row) if _clean_text(value)}


def _col(headers: dict[str, int], *names: str) -> int:
    for name in names:
        key = _header_key(name)
        if key in headers:
            return headers[key]
    raise ValueError(f"No se encontró ninguna columna: {', '.join(names)}")


def _scan_years(workbook: Any) -> list[int]:
    years: set[int] = set()
    for name in ("Base_Mes_Semana", "OMT", "Segundas Cx", "IPLH_TPLH_Real", "Costo", "CTC_DM", "CTC_Tienda", "Base_Qualtrics"):
        sheet = _required_sheet(workbook, name)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            number = _number(row[0] if row else None)
            if number is not None:
                years.add(int(number))
    return sorted(years)


def _latest_week(workbook: Any, year: int) -> int:
    specs = {
        "OMT": 3,
        "Segundas Cx": 3,
        "IPLH_TPLH_Real": 3,
        "Costo": 3,
        "CTC_DM": 1,
        "CTC_Tienda": 1,
        "Base_Qualtrics": 1,
    }
    weeks: list[int] = []
    for name, week_index in specs.items():
        for row in _required_sheet(workbook, name).iter_rows(min_row=2, values_only=True):
            row_year = _number(row[0] if row else None)
            week = _number(row[week_index] if len(row) > week_index else None)
            if row_year is not None and int(row_year) == year and week is not None:
                weeks.append(int(week))
    if not weeks:
        raise ValueError(f"No hay datos semanales para el año {year}")
    return max(weeks)


def _directory(workbook: Any) -> list[dict[str, str]]:
    sheet = _required_sheet(workbook, "Directorio")
    rows = sheet.iter_rows(values_only=True)
    next(rows, None)  # fila numérica heredada del archivo base
    headers = _header_map(next(rows, ()))
    indexes = {
        "ceco": _col(headers, "CC"),
        "tienda": _col(headers, "Tienda"),
        "region": _col(headers, "Región", "Region"),
        "estatus": _col(headers, "Estatus"),
        "dm": _col(headers, "DM"),
    }
    result: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in rows:
        ceco = normalize_ceco(row[indexes["ceco"]])
        if not ceco or ceco in seen:
            continue
        seen.add(ceco)
        result.append({
            "ceco": ceco,
            "tienda": _clean_text(row[indexes["tienda"]]),
            "region": _clean_text(row[indexes["region"]]),
            "dm": _clean_text(row[indexes["dm"]]),
            "estatus": _clean_text(row[indexes["estatus"]]),
        })
    return result


def _month_weeks(workbook: Any, year: int) -> tuple[dict[str, list[int]], dict[int, str]]:
    sheet = _required_sheet(workbook, "Base_Mes_Semana")
    grouped: dict[str, list[int]] = defaultdict(list)
    week_month: dict[int, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_year = _number(row[0] if row else None)
        week = _number(row[2] if len(row) > 2 else None)
        month = _clean_text(row[1] if len(row) > 1 else None).title()[:3]
        if row_year is None or int(row_year) != year or week is None or month not in MONTH_ORDER:
            continue
        week_int = int(week)
        if week_int not in grouped[month]:
            grouped[month].append(week_int)
        week_month[week_int] = month
    ordered = {month: sorted(grouped[month]) for month in MONTH_ORDER if grouped.get(month)}
    if not ordered:
        raise ValueError(f"Base_Mes_Semana no contiene un calendario válido para {year}")
    return ordered, week_month


class MetricsBuilder:
    def __init__(self, year: int, latest_week: int) -> None:
        self.year = year
        self.latest_week = latest_week
        self.records: dict[tuple[str, int], dict[str, Any]] = {}
        self.audits: list[SheetAudit] = []

    def record(self, ceco: Any, week: Any) -> dict[str, Any] | None:
        code = normalize_ceco(ceco)
        week_number = _number(week)
        if not code or week_number is None:
            return None
        week_int = int(week_number)
        if week_int < 1 or week_int > self.latest_week:
            return None
        return self.records.setdefault((code, week_int), {"ceco": code, "semana": week_int, "anio": self.year})

    def add_sheet(self, workbook: Any, name: str, loader: Callable[[tuple[Any, ...], dict[str, int], "MetricsBuilder"], tuple[str, int] | None]) -> None:
        sheet = _required_sheet(workbook, name)
        iterator = sheet.iter_rows(values_only=True)
        headers = _header_map(next(iterator, ()))
        rows_read = rows_in_year = rows_latest = duplicates = 0
        max_week: int | None = None
        seen: set[tuple[str, int]] = set()
        for row in iterator:
            rows_read += 1
            row_year = _number(row[0] if row else None)
            if row_year is None or int(row_year) != self.year:
                continue
            rows_in_year += 1
            key = loader(row, headers, self)
            if key is None:
                continue
            _, week = key
            max_week = week if max_week is None else max(max_week, week)
            if week == self.latest_week:
                rows_latest += 1
            if key in seen:
                duplicates += 1
            seen.add(key)
        self.audits.append(SheetAudit(name, rows_read, rows_in_year, max_week, rows_latest, duplicates))


def _set(record: dict[str, Any] | None, **values: Any) -> None:
    if record is None:
        return
    for key, value in values.items():
        clean = _rounded(value)
        if clean is not None:
            record[key] = clean


def _load_omt(row: tuple[Any, ...], h: dict[str, int], builder: MetricsBuilder) -> tuple[str, int] | None:
    ceco_i, week_i = _col(h, "CeCo"), _col(h, "Semana")
    record = builder.record(row[ceco_i], row[week_i])
    real, prior = _number(row[_col(h, "ADT Real")]), _number(row[_col(h, "ADT AA")])
    if real is not None and prior is not None:
        _set(record, omt=real - prior)
    return (record["ceco"], record["semana"]) if record else None


def _load_seg(row: tuple[Any, ...], h: dict[str, int], builder: MetricsBuilder) -> tuple[str, int] | None:
    record = builder.record(row[_col(h, "CeCo")], row[_col(h, "Semana")])
    value = _number(row[_col(h, "Segundas Cx")])
    if value is not None:
        _set(record, seg=value / 7)
    return (record["ceco"], record["semana"]) if record else None


def _load_labor(row: tuple[Any, ...], h: dict[str, int], builder: MetricsBuilder) -> tuple[str, int] | None:
    record = builder.record(row[_col(h, "CeCo")], row[_col(h, "Semana")])
    _set(record, iplh=row[_col(h, "IPLH")], tplh=row[_col(h, "TPLH")])
    return (record["ceco"], record["semana"]) if record else None


def _load_cost(row: tuple[Any, ...], h: dict[str, int], builder: MetricsBuilder) -> tuple[str, int] | None:
    record = builder.record(row[_col(h, "CeCo")], row[_col(h, "Semana")])
    _set(record, costo=row[_col(h, "Var Inventario")])
    return (record["ceco"], record["semana"]) if record else None


def _load_ctc_store(row: tuple[Any, ...], h: dict[str, int], builder: MetricsBuilder) -> tuple[str, int] | None:
    record = builder.record(row[_col(h, "Ceco", "CeCo")], row[_col(h, "Semana")])
    _set(record, ctc=row[_col(h, "Part FHW")])
    return (record["ceco"], record["semana"]) if record else None


def _load_qualtrics(row: tuple[Any, ...], h: dict[str, int], builder: MetricsBuilder) -> tuple[str, int] | None:
    ceco_i = h.get(_header_key("CeCo"))
    ceco = row[ceco_i] if ceco_i is not None and ceco_i < len(row) else None
    if normalize_ceco(ceco) is None:
        ceco = normalize_ceco(row[_col(h, "Centro de costos")], decimal_code=True)
    record = builder.record(ceco, row[_col(h, "Semana")])
    if record is None:
        return None
    _set(
        record,
        encuestas=row[_col(h, "Encuestas")],
        nps=row[_col(h, "NPS")],
        cx=parse_percent(row[_col(h, "Conexión", "Conexion")]),
        desempeno=parse_percent(row[_col(h, "Desempeño operacional", "Desempeno operacional")]),
        bebida=parse_percent(row[_col(h, "Sabor de la bebida")]),
    )
    return record["ceco"], record["semana"]


def _ctc_dm(workbook: Any, year: int, latest_week: int) -> tuple[list[dict[str, Any]], SheetAudit]:
    sheet = _required_sheet(workbook, "CTC_DM")
    iterator = sheet.iter_rows(values_only=True)
    headers = _header_map(next(iterator, ()))
    week_i, dm_i, value_i = _col(headers, "Semana"), _col(headers, "DM"), _col(headers, "Part FHW")
    result: list[dict[str, Any]] = []
    rows_read = rows_in_year = rows_latest = 0
    max_week: int | None = None
    for row in iterator:
        rows_read += 1
        row_year, week = _number(row[0]), _number(row[week_i])
        if row_year is None or int(row_year) != year or week is None:
            continue
        rows_in_year += 1
        week_int = int(week)
        max_week = week_int if max_week is None else max(max_week, week_int)
        if week_int > latest_week:
            continue
        if week_int == latest_week:
            rows_latest += 1
        item: dict[str, Any] = {"anio": year, "semana": week_int, "dm": _clean_text(row[dm_i])}
        value = _rounded(row[value_i])
        item["ctc"] = value
        result.append(item)
    return result, SheetAudit("CTC_DM", rows_read, rows_in_year, max_week, rows_latest)


def _atomic_write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as handle:
            handle.write(content)
        os.replace(temporary, path)
    except Exception:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass
        raise


def build_report(
    excel_path: str | Path,
    output_path: str | Path,
    audit_path: str | Path,
    *,
    year: int | None = None,
    expected_week: int | None = None,
) -> BuildResult:
    excel = Path(excel_path)
    if not excel.is_file():
        raise FileNotFoundError(f"No existe el Excel base: {excel}")
    workbook = load_workbook(excel, read_only=True, data_only=True)
    try:
        years = _scan_years(workbook)
        selected_year = year if year is not None else (max(years) if years else None)
        if selected_year is None or selected_year not in years:
            raise ValueError(f"Año no disponible. Años detectados: {years}")
        latest_week = _latest_week(workbook, selected_year)
        if expected_week is not None and latest_week < expected_week:
            raise ValueError(f"Se esperaba al menos la semana {expected_week}, pero el Excel llega a la {latest_week}")

        directory = _directory(workbook)
        month_weeks, week_month = _month_weeks(workbook, selected_year)
        default_month = week_month.get(latest_week, next(reversed(month_weeks)))

        builder = MetricsBuilder(selected_year, latest_week)
        builder.add_sheet(workbook, "OMT", _load_omt)
        builder.add_sheet(workbook, "Segundas Cx", _load_seg)
        builder.add_sheet(workbook, "IPLH_TPLH_Real", _load_labor)
        builder.add_sheet(workbook, "Costo", _load_cost)
        builder.add_sheet(workbook, "CTC_Tienda", _load_ctc_store)
        builder.add_sheet(workbook, "Base_Qualtrics", _load_qualtrics)
        ctc_dm, ctc_audit = _ctc_dm(workbook, selected_year, latest_week)

        metrics = sorted(builder.records.values(), key=lambda item: (int(item["ceco"]), item["semana"]))
        latest_coverage = defaultdict(int)
        for item in metrics:
            if item["semana"] == latest_week:
                for key in ("omt", "seg", "iplh", "tplh", "costo", "ctc", "encuestas", "nps", "cx", "desempeno", "bebida"):
                    if item.get(key) is not None:
                        latest_coverage[key] += 1

        payload = {
            "source": excel.name,
            "version": f"w{latest_week}-python-engine",
            "updatedToWeek": latest_week,
            "defaultMonth": default_month,
            "directory": directory,
            "monthWeeks": month_weeks,
            "ctcDM": ctc_dm,
            "metrics": metrics,
        }
        audit = {
            "status": "ok",
            "source": excel.name,
            "year": selected_year,
            "updatedToWeek": latest_week,
            "defaultMonth": default_month,
            "directoryRows": len(directory),
            "metricRows": len(metrics),
            "ctcDMRows": len(ctc_dm),
            "latestWeekCoverage": dict(sorted(latest_coverage.items())),
            "sheets": [asdict(item) for item in (*builder.audits, ctc_audit)],
        }
    finally:
        workbook.close()

    output = Path(output_path)
    audit_output = Path(audit_path)
    js = "// Generado automáticamente por engine/foco_engine. No editar manualmente.\nwindow.FOCO_DATA="
    js += json.dumps(payload, ensure_ascii=False, separators=(",", ":"), allow_nan=False) + ";\n"
    _atomic_write(output, js)
    _atomic_write(audit_output, json.dumps(audit, ensure_ascii=False, indent=2, allow_nan=False) + "\n")
    return BuildResult(selected_year, latest_week, default_month, len(directory), len(metrics), len(ctc_dm), str(output), str(audit_output))

