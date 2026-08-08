from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from .generator import _col, _ctc_dm, _directory, _locate_table, _month_weeks, _number, _rounded, normalize_ceco, parse_percent


METRICS = ("omt", "seg", "iplh", "tplh", "costo", "ctc", "encuestas", "nps", "cx", "desempeno", "bebida")


def load_data(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8")
    match = re.search(r"window\.FOCO_DATA=(.*);\s*$", text, re.S)
    if not match:
        raise ValueError("data.js no contiene window.FOCO_DATA válido")
    return json.loads(match.group(1))


def compare(expected: dict[tuple[str, int], Any], actual: dict[tuple[str, int], Any], metric: str) -> dict[str, Any]:
    missing = sorted(set(expected) - set(actual))
    extra = sorted(set(actual) - set(expected))
    mismatches = []
    for key in sorted(set(expected) & set(actual)):
        left, right = expected[key], actual[key]
        if abs(float(left) - float(right)) > 0.000001:
            mismatches.append({"key": f"{key[0]}|{key[1]}", "source": left, "output": right})
            if len(mismatches) >= 20:
                break
    return {
        "metric": metric,
        "sourceValues": len(expected),
        "outputValues": len(actual),
        "missing": len(missing),
        "extra": len(extra),
        "mismatches": mismatches,
        "sampleMissing": [f"{key[0]}|{key[1]}" for key in missing[:10]],
        "status": "ok" if not missing and not extra and not mismatches else "error",
    }


def scan_simple(
    workbook: Any,
    sheet_name: str,
    year: int,
    value_builders: dict[str, Callable[[tuple[Any, ...], dict[str, int]], Any]],
    *,
    ceco_names: tuple[str, ...] = ("CeCo", "Ceco"),
) -> dict[str, dict[tuple[str, int], Any]]:
    sheet = workbook[sheet_name]
    header_row, headers = _locate_table(sheet, ("Año", "Semana"))
    year_i, week_i, ceco_i = _col(headers, "Año"), _col(headers, "Semana"), _col(headers, *ceco_names)
    result = {metric: {} for metric in value_builders}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row_year, week = _number(row[year_i]), _number(row[week_i])
        code = normalize_ceco(row[ceco_i])
        if row_year is None or int(row_year) != year or week is None or not code:
            continue
        key = (code, int(week))
        for metric, builder in value_builders.items():
            value = _rounded(builder(row, headers))
            if value is not None:
                result[metric][key] = value
    return result


def scan_qualtrics(workbook: Any, year: int) -> dict[str, dict[tuple[str, int], Any]]:
    sheet = workbook["Base_Qualtrics"]
    header_row, headers = _locate_table(sheet, ("Año", "Semana", "Encuestas"))
    year_i, week_i = _col(headers, "Año"), _col(headers, "Semana")
    ceco_i = headers.get("ceco")
    accumulators: dict[tuple[str, int], dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row_year, week = _number(row[year_i]), _number(row[week_i])
        if row_year is None or int(row_year) != year or week is None:
            continue
        ceco = row[ceco_i] if ceco_i is not None else None
        code = normalize_ceco(ceco)
        if not code:
            code = normalize_ceco(row[_col(headers, "Centro de costos")], decimal_code=True)
        if not code:
            continue
        key = (code, int(week))
        item = accumulators.setdefault(key, {"encuestas": 0.0, "weighted": {}})
        surveys = _number(row[_col(headers, "Encuestas")])
        weight = surveys if surveys is not None and surveys > 0 else 1.0
        if surveys is not None:
            item["encuestas"] += surveys
        values = {
            "nps": _number(row[_col(headers, "NPS")]),
            "cx": parse_percent(row[_col(headers, "Conexión", "Conexion")]),
            "desempeno": parse_percent(row[_col(headers, "Desempeño operacional", "Desempeno operacional")]),
            "bebida": parse_percent(row[_col(headers, "Sabor de la bebida")]),
        }
        for metric, value in values.items():
            if value is None:
                continue
            total, total_weight = item["weighted"].get(metric, (0.0, 0.0))
            item["weighted"][metric] = (total + value * weight, total_weight + weight)
    result = {metric: {} for metric in ("encuestas", "nps", "cx", "desempeno", "bebida")}
    for key, item in accumulators.items():
        if item["encuestas"]:
            result["encuestas"][key] = _rounded(item["encuestas"])
        for metric, (total, weight) in item["weighted"].items():
            if weight:
                result[metric][key] = _rounded(total / weight)
    return result


def validate(excel_path: Path, data_path: Path, audit_path: Path, report_path: Path) -> dict[str, Any]:
    data = load_data(data_path)
    audit = json.loads(audit_path.read_text(encoding="utf-8"))
    source_sha = hashlib.sha256(excel_path.read_bytes()).hexdigest()
    if data.get("sourceSha256") != source_sha or audit.get("sourceSha256") != source_sha:
        raise ValueError("data.js o la auditoría no corresponden al Excel actual (SHA-256 diferente)")
    year = int(data["metrics"][0]["anio"] if data.get("metrics") else audit["year"])
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    structural_checks: list[dict[str, Any]] = []
    try:
        required = {"Directorio", "Base_Mes_Semana", "OMT", "Segundas Cx", "IPLH_TPLH_Real", "Costo", "CTC_DM", "CTC_Tienda", "Base_Qualtrics"}
        missing_sheets = sorted(required - set(workbook.sheetnames))
        if missing_sheets:
            raise ValueError(f"Faltan pestañas: {', '.join(missing_sheets)}")
        expected: dict[str, dict[tuple[str, int], Any]] = {}
        expected.update(scan_simple(workbook, "OMT", year, {
            "omt": lambda row, h: (_number(row[_col(h, "ADT Real")]) - _number(row[_col(h, "ADT AA")]))
            if _number(row[_col(h, "ADT Real")]) is not None and _number(row[_col(h, "ADT AA")]) is not None else None,
        }))
        expected.update(scan_simple(workbook, "Segundas Cx", year, {
            "seg": lambda row, h: _number(row[_col(h, "Segundas Cx")]) / 7 if _number(row[_col(h, "Segundas Cx")]) is not None else None,
        }))
        expected.update(scan_simple(workbook, "IPLH_TPLH_Real", year, {
            "iplh": lambda row, h: row[_col(h, "IPLH")], "tplh": lambda row, h: row[_col(h, "TPLH")],
        }))
        expected.update(scan_simple(workbook, "Costo", year, {"costo": lambda row, h: row[_col(h, "Var Inventario")]}))
        expected.update(scan_simple(workbook, "CTC_Tienda", year, {"ctc": lambda row, h: row[_col(h, "Part FHW")]}))
        expected.update(scan_qualtrics(workbook, year))
        expected_directory = _directory(workbook)
        source_directory = {item["ceco"]: item for item in expected_directory}
        output_directory = {str(item["ceco"]): item for item in data.get("directory", [])}
        directory_ok = source_directory == output_directory
        structural_checks.append({"name": "Directorio", "sourceRows": len(source_directory), "outputRows": len(output_directory), "status": "ok" if directory_ok else "error"})
        expected_calendar, _ = _month_weeks(workbook, year)
        calendar_ok = expected_calendar == data.get("monthWeeks")
        structural_checks.append({"name": "Base_Mes_Semana", "sourceMonths": len(expected_calendar), "outputMonths": len(data.get("monthWeeks", {})), "status": "ok" if calendar_ok else "error"})
        expected_ctc_dm, _ = _ctc_dm(workbook, year, int(data["updatedToWeek"]))
        ctc_dm_ok = expected_ctc_dm == data.get("ctcDM")
        structural_checks.append({"name": "CTC_DM", "sourceRows": len(expected_ctc_dm), "outputRows": len(data.get("ctcDM", [])), "status": "ok" if ctc_dm_ok else "error"})
    finally:
        workbook.close()
    records = {(str(item["ceco"]), int(item["semana"])): item for item in data["metrics"]}
    actual = {metric: {key: item[metric] for key, item in records.items() if item.get(metric) is not None} for metric in METRICS}
    checks = [compare(expected[metric], actual[metric], metric) for metric in METRICS]
    latest = int(data["updatedToWeek"])
    latest_counts = {metric: sum(1 for (code, week) in actual[metric] if week == latest) for metric in METRICS}
    report = {
        "status": "ok" if all(check["status"] == "ok" for check in checks + structural_checks) else "error",
        "source": excel_path.name,
        "sourceSha256": source_sha,
        "year": year,
        "updatedToWeek": latest,
        "latestWeekCoverage": latest_counts,
        "structuralChecks": structural_checks,
        "checks": checks,
    }
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    if report["status"] != "ok":
        failed = ", ".join([check["metric"] for check in checks if check["status"] != "ok"] + [check["name"] for check in structural_checks if check["status"] != "ok"])
        raise ValueError(f"La reconciliación fuente→data.js falló en: {failed}")
    if any(latest_counts.get(metric, 0) == 0 for metric in ("omt", "iplh", "tplh")):
        raise ValueError(f"OMT o IPLH/TPLH no contienen la semana {latest}")
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Reconciliación independiente Excel → data.js")
    parser.add_argument("--excel", required=True)
    parser.add_argument("--data", required=True)
    parser.add_argument("--audit", required=True)
    parser.add_argument("--report", default="VALIDACION_DATOS.json")
    args = parser.parse_args()
    try:
        report = validate(Path(args.excel), Path(args.data), Path(args.audit), Path(args.report))
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1
    print(json.dumps({"status": report["status"], "updatedToWeek": report["updatedToWeek"], "latestWeekCoverage": report["latestWeekCoverage"]}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
